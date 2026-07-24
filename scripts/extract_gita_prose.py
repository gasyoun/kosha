#!/usr/bin/env python
"""H1493 — vendor the Gītā.xlsm *Prose* sheet as an interlinear paraphrase TSV.

The Prose sheet is a running interlinear paraphrase (form + parenthetical gloss),
not the word-by-word Combined master. Multi-row blocks share a verse ref only on
the first line; continuation rows have a blank ref. Verse labels may be single
(`1.12`) or ranges (`1.4-6`, `1.15-16`).

Usage:
  python scripts/extract_gita_prose.py
  python scripts/extract_gita_prose.py --xlsm path/to/Gita.xlsm
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSM = ROOT.parent / "SanskritGrammar" / "Concordance" / "Gita.xlsm"
OUT = ROOT / "data" / "gita" / "gita_prose.tsv"
OUT_JS = ROOT / "reading" / "data" / "gita_prose.js"

# 1.12  |  1.4-6  |  1.15-16
VERSE_RE = re.compile(
    r"^\s*(\d+)\s*\.\s*(\d+)(?:\s*[-–—]\s*(\d+))?\s*$"
)


def cell(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\xa0", " ").replace("\t", " ").replace("\n", " ").strip()


def expand_verse_keys(label: str) -> list[str]:
    """'1.4-6' → ['1.4','1.5','1.6']; '1.12' → ['1.12']."""
    m = VERSE_RE.match(label or "")
    if not m:
        return [label] if label else []
    ch, a, b = int(m.group(1)), int(m.group(2)), m.group(3)
    end = int(b) if b else a
    if end < a:
        a, end = end, a
    return [f"{ch}.{n}" for n in range(a, end + 1)]


def extract(xlsm: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsm), read_only=True, data_only=True)
    if "Prose" not in wb.sheetnames:
        sys.exit(f"no Prose sheet in {xlsm}; sheets={wb.sheetnames}")
    ws = wb["Prose"]

    blocks: list[dict] = []
    cur_label = ""
    cur_lines: list[str] = []
    line_n = 0

    def flush():
        nonlocal cur_label, cur_lines, line_n
        if not cur_label or not cur_lines:
            cur_label, cur_lines, line_n = "", [], 0
            return
        text = " ".join(cur_lines).strip()
        text = re.sub(r"\s+", " ", text)
        keys = expand_verse_keys(cur_label)
        blocks.append(
            {
                "verse_label": cur_label,
                "verse_keys": "|".join(keys),
                "n_lines": str(line_n),
                "text": text,
            }
        )
        cur_label, cur_lines, line_n = "", [], 0

    for row in ws.iter_rows(min_row=1, values_only=True):
        ref = cell(row[0] if row else None)
        txt = cell(row[1] if row and len(row) > 1 else None)
        if not txt and not ref:
            continue
        if ref:
            flush()
            cur_label = ref
            line_n = 0
        if not cur_label:
            # orphan continuation before any verse label — skip
            continue
        if txt:
            line_n += 1
            cur_lines.append(txt)
    flush()
    wb.close()
    return blocks


def write_tsv(blocks: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    cols = ["verse_label", "verse_keys", "n_lines", "text"]
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, delimiter="\t", lineterminator="\n")
        w.writeheader()
        for b in blocks:
            w.writerow({c: b.get(c, "") for c in cols})


def write_js(blocks: list[dict], path: Path) -> None:
    """window.GITA_PROSE[verse] = text  (ranges expanded to every member)."""
    import json

    by_verse: dict[str, str] = {}
    for b in blocks:
        keys = (b.get("verse_keys") or "").split("|")
        text = b.get("text") or ""
        for k in keys:
            if k:
                # last write wins if overlapping ranges (should not happen)
                by_verse[k] = text
    path.parent.mkdir(parents=True, exist_ok=True)
    body = json.dumps(by_verse, ensure_ascii=False, indent=1, sort_keys=True)
    with path.open("w", encoding="utf-8", newline="\n") as f:
        f.write("window.GITA_PROSE = ")
        f.write(body)
        f.write(";\n")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsm", default=str(DEFAULT_XLSM))
    ap.add_argument("--out", default=str(OUT))
    ap.add_argument("--js", default=str(OUT_JS))
    args = ap.parse_args()
    xlsm = Path(args.xlsm)
    if not xlsm.exists():
        sys.exit(f"MISSING xlsm: {xlsm}")
    blocks = extract(xlsm)
    write_tsv(blocks, Path(args.out))
    write_js(blocks, Path(args.js))
    n_keys = sum(len((b.get("verse_keys") or "").split("|")) for b in blocks)
    print(f"wrote {args.out} — {len(blocks)} prose blocks, {n_keys} verse keys")
    print(f"wrote {args.js}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
