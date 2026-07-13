#!/usr/bin/env python
"""H875 / roadmap W5 — Gītā etymology notes.

The `Combined` master drops the `Grammar` sheet's etymology-notes column
(col AG, e.g. `traditionally: pu-tra…`, `uttama – uppermost…`). This extracts
those notes, aligned to each word by (verse, word-index), into a vendored TSV.

Public/MIT, credit Dr. Mārcis Gasūns. Usage:
    python scripts/extract_gita_etymology.py --xlsm <path>
"""
import argparse
import csv
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSM = ROOT.parent / "SanskritGrammar" / "Concordance" / "Gita.xlsm"
OUT = ROOT / "data" / "gita" / "gita_etymology.tsv"

# Grammar sheet columns (0-indexed)
G = {"form": 3, "vref": 8, "widx": 9, "lemma": 13, "root": 14, "etym": 32}


def cell(r, k):
    v = r[G[k]] if G[k] < len(r) else None
    return "" if v is None else str(v).strip().replace("\t", " ").replace("\n", " ")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsm", default=str(DEFAULT_XLSM))
    ap.add_argument("--max-row", type=int, default=9200)
    args = ap.parse_args()
    import openpyxl
    ws = openpyxl.load_workbook(args.xlsm, read_only=True, data_only=True)["Grammar"]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with open(OUT, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", lineterminator="\n")
        w.writerow(["verse", "widx", "form", "lemma", "root", "etymology"])
        for r in ws.iter_rows(min_row=2, max_row=args.max_row, values_only=True):
            vref = cell(r, "vref")
            if not re.match(r"^\d+\.\d+", vref):
                continue
            etym = cell(r, "etym")
            if not etym:
                continue
            w.writerow([vref, cell(r, "widx"), cell(r, "form"),
                        cell(r, "lemma"), cell(r, "root"), etym])
            n += 1
    print(f"wrote {OUT} — {n} etymology notes")


if __name__ == "__main__":
    main()
