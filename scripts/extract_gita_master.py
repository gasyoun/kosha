#!/usr/bin/env python
"""W0 — vendor the canonical Gītā gold master dataset (H848 roadmap).

Extracts the author's clean labelled `Combined` sheet from
SanskritGrammar/Concordance/Gita.xlsm (9,091 analysed words, all 18 adhyāyas,
24 columns) into a committed TSV — the master every Gītā-gold workstream derives
from. Per MG's 13-07-2026 rulings: the garbled private-use-encoded Russian
*transliteration* column is dropped (the clean Cyrillic *gloss* is kept); the
derived dataset is MIT, public tier.

The xlsm is a local-only artifact (not committed in SanskritGrammar), so this
one-time extraction vendors the data into kosha; openpyxl reads it directly.

Usage: python scripts/extract_gita_master.py --xlsm <path>
"""
import argparse
import csv
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSM = ROOT.parent / "SanskritGrammar" / "Concordance" / "Gita.xlsm"
OUT = ROOT / "data" / "gita" / "gita_gold_master.tsv"

# Combined-sheet columns (0-indexed) -> output field. Col 6 (garbled Russian
# transliteration, private-use font) is intentionally omitted per MG ruling.
FIELDS = [
    (0, "verse"), (3, "lemma"), (4, "devanagari"), (5, "iast"),
    (7, "form_type"), (8, "code"), (9, "tense"), (10, "pada"), (11, "vclass"),
    (12, "root"), (13, "root_tr"), (14, "prefix"), (15, "stem_end"),
    (16, "gender"), (17, "compound"), (18, "mark"), (19, "rule"),
    (20, "sandhi"), (21, "verse_iast"), (22, "gloss_en"), (23, "gloss_ru"),
]
VERSE_RE = re.compile(r"^\d+\.\d+")


def cell(r, j):
    v = r[j] if j < len(r) else None
    return "" if v is None else str(v).strip().replace("\t", " ").replace("\n", " ")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsm", default=str(DEFAULT_XLSM))
    ap.add_argument("--max-row", type=int, default=9300)
    args = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsm, read_only=True, data_only=True)
    ws = wb["Combined"]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    chaps = {}
    with open(OUT, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", lineterminator="\n")
        w.writerow([name for _, name in FIELDS])
        for r in ws.iter_rows(min_row=2, max_row=args.max_row, values_only=True):
            verse = cell(r, 0)
            if not VERSE_RE.match(verse):
                continue
            w.writerow([cell(r, j) for j, _ in FIELDS])
            n += 1
            ch = int(verse.split(".")[0])
            chaps[ch] = chaps.get(ch, 0) + 1
    print(f"wrote {OUT} — {n} words, chapters {sorted(chaps)} counts {chaps}")


if __name__ == "__main__":
    main()
