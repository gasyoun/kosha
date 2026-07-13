#!/usr/bin/env python
"""H873 / roadmap W3 — Gītā gold morphology + compound dataset.

Parses the hand-curated morphology shorthand in the `Grammar` sheet of
SanskritGrammar/Concordance/Gita.xlsm (column AB, e.g. `1n.1 m.`, `Perf. P 1v.1`,
`PP 1n.3 m.`) into structured gold fields — case·number·gender for nominals,
person·number·tense·voice for finite verbs, non-finite/derivation tags, and the
compound type (TP/BV/DV/KD) — using the workbook's own `Abbreviations` legend
(embedded below verbatim). Vendors the result as a committed TSV.

Public/MIT, credit Dr. Mārcis Gasūns (roadmap §7). The xlsm is local-only, so
this one-time extraction vendors the data into kosha.

Usage: python scripts/extract_gita_morphology.py --xlsm <path>
"""
import argparse
import csv
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSM = ROOT.parent / "SanskritGrammar" / "Concordance" / "Gita.xlsm"
OUT = ROOT / "data" / "gita" / "gita_morphology_gold.tsv"

# --- decode maps, verbatim from the Gita.xlsm `Abbreviations` sheet ---
CASE = {1: "nom", 2: "acc", 3: "ins", 4: "dat", 5: "abl", 6: "gen", 7: "loc", 8: "voc"}
NUM = {1: "sg", 2: "du", 3: "pl"}
PERSON = {1: "3", 2: "2", 3: "1"}  # 1v=3rd, 2v=2nd, 3v=1st person
TENSE = {"Praes.": "present", "Pot.": "optative", "Imperat.": "imperative",
         "Imperf.": "imperfect", "Aor.": "aorist", "Perf.": "perfect",
         "Ben.": "benedictive", "Fut.p.": "periphrastic-future", "Fut.": "future",
         "Cond.": "conditional"}
DERIV = {"caus.": "causative", "des.": "desiderative", "intens.": "intensive",
         "denom.": "denominative", "pass.": "passive"}
NONFIN = {"inf.": "infinitive", "absol.": "absolutive", "PF": "future-participle",
          "PP": "past-participle", "PPr": "present-participle"}
GENDER = {"m": "m", "f": "f", "n": "n"}
VOICE = {"P": "parasmaipada", "Ā": "atmanepada"}
COMPOUND = {"BV": "bahuvrihi", "DV": "dvandva", "KD": "karmadharaya", "TP": "tatpurusa"}

NCASE = re.compile(r"^([1-8])n\.([1-3])$")   # nominal: <case>n.<number>
VPERS = re.compile(r"^([1-3])v\.([1-3])$")   # verb:    <person>v.<number>

# Grammar sheet columns (0-indexed)
G = {"form": 3, "chapter": 7, "vref": 8, "widx": 9, "lemma": 13, "root": 14,
     "gender_col": 16, "compound": 18, "morph": 27}


def parse_morph(ab, gender_col, compound_col):
    """AB shorthand -> structured dict."""
    out = {"pos": "", "case": "", "number": "", "gender": "", "person": "",
           "tense": "", "voice": "", "nonfinite": "", "derivation": "",
           "compound": COMPOUND.get(compound_col.strip(), compound_col.strip())}
    if gender_col and gender_col.strip() in GENDER:
        out["gender"] = GENDER[gender_col.strip()]
    for tok in (ab or "").replace(",", " ").split():
        t = tok.strip().strip(".") + ("." if tok.strip().endswith(".") else "")
        t = tok.strip()
        m = NCASE.match(t)
        if m:
            out["pos"] = out["pos"] or "nominal"
            out["case"] = CASE[int(m.group(1))]; out["number"] = NUM[int(m.group(2))]; continue
        m = VPERS.match(t)
        if m:
            out["pos"] = "verb"
            out["person"] = PERSON[int(m.group(1))]; out["number"] = NUM[int(m.group(2))]; continue
        if t in TENSE: out["tense"] = TENSE[t]; out["pos"] = out["pos"] or "verb"; continue
        if t in DERIV: out["derivation"] = DERIV[t]; continue
        if t in NONFIN: out["nonfinite"] = NONFIN[t]; out["pos"] = "participle" if "participle" in NONFIN[t] else out["pos"]; continue
        if t in VOICE: out["voice"] = VOICE[t]; continue
        if t.rstrip(".") in ("m", "f", "n") and not out["gender"]:
            out["gender"] = t.rstrip("."); continue
        if t.rstrip(".") in ("av", "sn"):
            out["pos"] = out["pos"] or ("indeclinable" if t.rstrip(".") == "av" else "pronoun")
    if not out["pos"]:
        out["pos"] = "nominal" if out["case"] else ""
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsm", default=str(DEFAULT_XLSM))
    ap.add_argument("--max-row", type=int, default=9200)
    args = ap.parse_args()
    import openpyxl
    ws = openpyxl.load_workbook(args.xlsm, read_only=True, data_only=True)["Grammar"]

    cols = ["verse", "widx", "form", "lemma", "root", "pos", "case", "number",
            "gender", "person", "tense", "voice", "nonfinite", "derivation",
            "compound", "raw_morph"]
    OUT.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with open(OUT, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", lineterminator="\n"); w.writerow(cols)
        for r in ws.iter_rows(min_row=2, max_row=args.max_row, values_only=True):
            def c(k):
                v = r[G[k]] if G[k] < len(r) else None
                return "" if v is None else str(v).strip()
            vref = c("vref")
            if not re.match(r"^\d+\.\d+", vref):
                continue
            ab = c("morph")
            p = parse_morph(ab, c("gender_col"), c("compound"))
            w.writerow([vref, c("widx"), c("form"), c("lemma"), c("root"),
                        p["pos"], p["case"], p["number"], p["gender"], p["person"],
                        p["tense"], p["voice"], p["nonfinite"], p["derivation"],
                        p["compound"], ab])
            n += 1
    print(f"wrote {OUT} — {n} words parsed")


if __name__ == "__main__":
    main()
